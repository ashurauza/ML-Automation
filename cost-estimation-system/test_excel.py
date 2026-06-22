import sys
sys.path.append('backend')
from services.excel_exporter import generate_complex_excel

# Overwrite generate_complex_excel to test the fix
import io
from openpyxl import load_workbook
import os
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname('backend/services/excel_exporter.py')))
TEMPLATE_PATH = os.path.join(BASE_DIR, "Costing", "Quotation+BOM all of package PC4100.xlsx")

def generate_complex_excel_fixed(estimation) -> io.BytesIO:
    wb = load_workbook(TEMPLATE_PATH)
    costs = {
        "Raw Material Cost": getattr(estimation, "raw_material_cost", 0.0) or 0.0,
        "Total Cost": getattr(estimation, "total_cost", 0.0) or 0.0,
    }
    ws_ml = wb.create_sheet("AI Estimation Results", 0)
    
    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)

    ws_ml['A1'] = "Cost Estimation Summary"
    ws_ml['A1'].font = header_font
    
    ws_ml['A4'] = "Cost Component"
    ws_ml['A4'].font = bold_font
    ws_ml['B4'] = "Estimated Value (INR)"
    ws_ml['B4'].font = bold_font
    
    row = 5
    for component, value in costs.items():
        ws_ml.cell(row=row, column=1, value=component)
        cell = ws_ml.cell(row=row, column=2, value=value)
        if component == "Total Cost":
            ws_ml.cell(row=row, column=1).font = bold_font
            cell.font = bold_font
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

class DummyEst:
    filename = 'test.pdf'
    raw_material_cost = 100.0
    total_cost = 100.0

generate_complex_excel_fixed(DummyEst())
print("Success")
