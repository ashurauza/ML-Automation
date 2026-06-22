import io
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Base directory
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
TEMPLATE_PATH = os.path.join(BASE_DIR, "Costing", "Quotation+BOM all of package PC4100.xlsx")

def generate_complex_excel(estimation) -> io.BytesIO:
    """
    Generates a complex multi-sheet Excel file by cloning a highly detailed template
    and injecting dynamic values from the estimation model into a new summary sheet.
    """
    # Load the template workbook
    wb = load_workbook(TEMPLATE_PATH)
    
    # Extract ML predictions from the estimation object
    costs = {
        "Raw Material Cost": getattr(estimation, "raw_material_cost", 0.0) or 0.0,
        "Machining Cost": getattr(estimation, "machining_cost", 0.0) or 0.0,
        "Manpower Cost": getattr(estimation, "manpower_cost", 0.0) or 0.0,
        "Coating Cost": getattr(estimation, "coating_cost", 0.0) or 0.0,
        "Overhead Cost": getattr(estimation, "overhead_cost", 0.0) or 0.0,
        "Logistics Cost": getattr(estimation, "logistics_cost", 0.0) or 0.0,
        "Profit Margin": getattr(estimation, "profit_margin", 0.0) or 0.0,
        "Total Cost": getattr(estimation, "total_cost", 0.0) or 0.0,
    }

    # Add a dedicated sheet at the front for the ML output to prevent 
    # breaking the complex merged cells / formulas of the Rig Adaptor template.
    ws_ml = wb.create_sheet("AI Estimation Results", 0)
    
    # Set up styling and headers
    ws_ml.column_dimensions['A'].width = 30
    ws_ml.column_dimensions['B'].width = 20
    
    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)

    ws_ml['A1'] = "Cost Estimation Summary"
    ws_ml['A1'].font = header_font
    ws_ml['A2'] = f"File: {estimation.filename}"
    
    ws_ml['A4'] = "Cost Component"
    ws_ml['A4'].font = bold_font
    ws_ml['B4'] = "Estimated Value (INR)"
    ws_ml['B4'].font = bold_font
    
    # Write the data
    row = 5
    for component, value in costs.items():
        ws_ml.cell(row=row, column=1, value=component)
        cell = ws_ml.cell(row=row, column=2, value=value)
        cell.number_format = '₹#,##0.00'
        if component == "Total Cost":
            ws_ml.cell(row=row, column=1).font = bold_font
            cell.font = bold_font
        row += 1

    # Extract additional ML details if available
    if hasattr(estimation, "diagram_analysis") and estimation.diagram_analysis:
        ml_details = estimation.diagram_analysis.get("ml_details", {})
        if ml_details:
            row += 2
            ws_ml.cell(row=row, column=1, value="AI Confidence Metrics").font = bold_font
            row += 1
            ws_ml.cell(row=row, column=1, value="Lower Bound")
            ws_ml.cell(row=row, column=2, value=ml_details.get("confidence_lower", 0.0)).number_format = '₹#,##0.00'
            row += 1
            ws_ml.cell(row=row, column=1, value="Upper Bound")
            ws_ml.cell(row=row, column=2, value=ml_details.get("confidence_upper", 0.0)).number_format = '₹#,##0.00'

    # Save to a memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
